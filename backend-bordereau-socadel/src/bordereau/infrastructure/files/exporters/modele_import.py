"""Génération du classeur vierge de saisie. Implémente `GenerateurModeleImport`.

Les en-têtes sont exactement ceux que le lecteur d'import reconnaît : le
superviseur ne peut pas produire, avec ce modèle, un fichier que l'application
refuserait ensuite pour cause de colonnes inattendues.

L'aide de saisie est portée par des **commentaires de cellule** sur l'en-tête,
et non par une ligne de texte sous celui-ci : une telle ligne serait relue à
l'import comme une donnée sans contrat, et le superviseur verrait à chaque fois
une ligne rejetée fantôme.
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from ....domain.enums import Responsable, StatutCollecte

#: (en-tête, largeur en caractères, aide portée par le commentaire de cellule)
COLONNES = (
    ("SERVICE_NO", 16, "Obligatoire — numéro de contrat du client."),
    ("NOMS", 32, "Nom du client tel qu'il figure au bordereau."),
    ("REF_GEO", 22, "Référence géographique, ex. 807-09-01-994-00-001."),
    ("ITINERAIRE", 12, "Code de l'itinéraire parcouru, ex. 130387."),
    ("METER_NO", 18, "Numéro de compteur."),
    ("NUMERO_TELEPHONE", 18, "Numéro WhatsApp relevé, ex. +237694174768."),
    (
        "STATUT",
        16,
        "ABONNE, NON_ABONNE, ABSENT, INJOIGNABLE, REFUS ou DOUBLON.\n"
        "Un statut ABONNE exige le numéro relevé.",
    ),
    ("RESPONSABLE", 14, "TERRAIN, CHATBOT, CSC ou AUTRES."),
    ("OBSERVATION", 34, "Remarque libre de l'agent."),
)

BLEU_SOCADEL = "FF1A76B9"
GRIS_BORDURE = "FFCBD5E1"


class GenerateurModeleXlsx:
    """Produit le modèle `.xlsx` distribué aux agents."""

    #: Nombre de lignes couvertes par les listes déroulantes.
    LIGNES_PREPAREES = 500

    def generer(self) -> bytes:
        classeur = Workbook()
        feuille = classeur.active
        feuille.title = "Bordereau"

        bordure = Border(
            left=Side(style="thin", color=GRIS_BORDURE),
            right=Side(style="thin", color=GRIS_BORDURE),
            top=Side(style="thin", color=GRIS_BORDURE),
            bottom=Side(style="thin", color=GRIS_BORDURE),
        )

        for index, (entete, largeur, aide) in enumerate(COLONNES, start=1):
            cellule = feuille.cell(row=1, column=index, value=entete)
            cellule.font = Font(bold=True, color="FFFFFFFF", size=10)
            cellule.fill = PatternFill("solid", fgColor=BLEU_SOCADEL)
            cellule.alignment = Alignment(horizontal="center", vertical="center")
            cellule.border = bordure
            # L'aide reste au plus près de la saisie sans polluer les données.
            cellule.comment = Comment(aide, "Bordereau SOCADEL", height=90, width=260)

            feuille.column_dimensions[get_column_letter(index)].width = largeur

        # Les données commencent en ligne 2, immédiatement sous l'en-tête.
        feuille.freeze_panes = "A2"
        feuille.row_dimensions[1].height = 22

        self._ajouter_listes_deroulantes(feuille)

        tampon = io.BytesIO()
        classeur.save(tampon)
        return tampon.getvalue()

    def _ajouter_listes_deroulantes(self, feuille) -> None:
        """Contraint les colonnes à vocabulaire fermé.

        Cela supprime la source d'anomalies la plus fréquente à l'import : un
        statut écrit « ok » ou « abonné » au lieu de la valeur attendue.
        """
        derniere_ligne = 1 + self.LIGNES_PREPAREES

        validations = (
            ("G", [s.value for s in StatutCollecte], "Statut invalide"),
            ("H", [r.value for r in Responsable], "Responsable invalide"),
        )

        for colonne, valeurs, message in validations:
            validation = DataValidation(
                type="list",
                formula1=f'"{",".join(valeurs)}"',
                allow_blank=True,
                showErrorMessage=True,
                errorTitle="Valeur non autorisée",
                error=f"{message}. Choisissez une valeur dans la liste.",
            )
            feuille.add_data_validation(validation)
            validation.add(f"{colonne}2:{colonne}{derniere_ligne}")
