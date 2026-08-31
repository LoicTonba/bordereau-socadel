"""Le bordereau de terrain en classeur, calqué sur `bordereau.xlsx / Feuil3`.

C'est le document que l'agent emporte. Il existe en deux formats, PDF et
Excel, et les deux portent la même maquette : un bandeau par itinéraire, la
grille de saisie, et la ligne de signature qui authentifie le retour de
tournée.

Pourquoi un Excel alors que le PDF suffit à imprimer : un superviseur qui
prépare une tournée hors application, ou qui doit ajouter des clients à la
main, travaille dans un tableur. Le classeur lui donne ce point d'entrée sans
qu'il ait à recomposer la maquette.

Deux choses qu'Excel ne sait pas faire et qu'on ne fait donc pas semblant de
faire. Il n'a pas de vrai filigrane : le logo est posé en tête de feuille, où
il s'imprime, plutôt que simulé par une image de fond qui se décalerait au
premier ajout de ligne. Et il n'a pas de zone protégée fiable : les colonnes
pré-remplies restent modifiables, la vérification au référentiel étant de
toute façon l'arbitre final.
"""

from __future__ import annotations

import io
from collections.abc import Sequence
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ImageXl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

TITRE_CAMPAGNE = "CAMPAGNE DE COLLECTE DE NUMERO WHATSAPP"

#: Colonnes de la feuille 3, plus la colonne WhatsApp. RAPPORT et N° WHATSAPP
#: sont les deux seules que l'agent remplit sur le terrain.
COLONNES = (
    ("REF GEO", 24),
    ("METER_NO", 18),
    ("NOMS", 32),
    ("CONTRAT", 14),
    ("RAPPORT", 12),
    ("N° WHATSAPP", 20),
)

BLEU = "FF1A76B9"
BLEU_SOMBRE = "FF1F5FA0"
BLEU_CLAIR = "FFEFF7FD"
BLANC = "FFFFFFFF"
GRIS_BORDURE = "FFCBD5E1"
GRIS_TEXTE = "FF334155"
GRIS_DOUX = "FF64748B"
CREME = "FFFAFCFE"

#: Lignes vierges laissées en fin de tournée : l'agent y note les clients
#: rencontrés qui ne figurent pas encore au référentiel.
LIGNES_VIERGES = 3

#: Même fichier que celui du filigrane des PDF : une seule marque, un seul
#: emplacement, et rien à resynchroniser.
LOGO = Path(__file__).parent / "assets" / "logo-socadel.jpg"


@dataclass(frozen=True, slots=True)
class LigneTerrain:
    """Un client à démarcher, réduit à ce que le papier montre."""

    ref_geo: str
    numero_compteur: str
    nom: str
    service_no: str


@dataclass(frozen=True, slots=True)
class BlocTerrain:
    """Une tournée : son code, son libellé, ses clients."""

    code: int
    libelle: str
    lignes: Sequence[LigneTerrain]


#: Tournée d'exemple, reprise du classeur source. Elle sert de mode d'emploi :
#: l'agent voit ce qu'on attend de lui avant même d'avoir reçu sa vraie
#: affectation.
EXEMPLE = BlocTerrain(
    code=125369,
    libelle="Exemple, à remplacer par votre tournée",
    lignes=[
        LigneTerrain("838-01-01-641-00-01", "021750196246", "BILOA DAMARIS", "203299981"),
        LigneTerrain("838-01-01-641-00-02", "021750196246", "NGONO ALBERTINE", "203299982"),
        LigneTerrain("838-01-01-641-00-03", "021750196246", "MBALLA JEAN PIERRE", "203299983"),
        LigneTerrain("838-01-01-641-00-04", "021750196246", "TCHOUMI ALAIN", "203299984"),
        LigneTerrain("838-01-01-641-00-05", "021750196246", "ABDOUL AZIZ OUMAROU", "203299985"),
    ],
)


class GenerateurModeleTerrainXlsx:
    """Produit le classeur de terrain, vierge ou renseigné."""

    def generer(
        self,
        blocs: Sequence[BlocTerrain] = (),
        *,
        nom_agent: str = "",
        date_travail: str = "",
    ) -> bytes:
        """Rend le classeur. Sans bloc, la tournée d'exemple sert de modèle."""
        classeur = Workbook()
        feuille = classeur.active
        feuille.title = "Bordereau terrain"

        _mettre_en_page(feuille)
        ligne = _entete_document(feuille, nom_agent, date_travail)

        for bloc in blocs or (EXEMPLE,):
            ligne = _bloc(feuille, bloc, ligne)

        ligne = _legende(feuille, ligne + 1)
        _signature(feuille, ligne + 1)

        tampon = io.BytesIO()
        classeur.save(tampon)
        return tampon.getvalue()


# --- Mise en page ----------------------------------------------------------


def _bordure() -> Border:
    trait = Side(style="thin", color=GRIS_BORDURE)
    return Border(left=trait, right=trait, top=trait, bottom=trait)


def _mettre_en_page(feuille: Worksheet) -> None:
    """Largeurs, orientation, et répétition de l'en-tête à l'impression."""
    for index, (_, largeur) in enumerate(COLONNES, start=1):
        feuille.column_dimensions[get_column_letter(index)].width = largeur

    feuille.page_setup.orientation = "portrait"
    feuille.page_setup.fitToWidth = 1
    feuille.page_setup.fitToHeight = 0
    feuille.sheet_properties.pageSetUpPr.fitToPage = True
    feuille.page_margins.left = feuille.page_margins.right = 0.4
    feuille.print_options.horizontalCentered = True


def _fusionner(feuille: Worksheet, ligne: int) -> str:
    """Fusionne la ligne sur toute la largeur et renvoie sa première cellule."""
    derniere = get_column_letter(len(COLONNES))
    feuille.merge_cells(f"A{ligne}:{derniere}{ligne}")
    return f"A{ligne}"


def _entete_document(feuille: Worksheet, nom_agent: str, date_travail: str) -> int:
    """Titre centré, marque, et la ligne d'identification de la tournée."""
    cellule = feuille[_fusionner(feuille, 1)]
    cellule.value = TITRE_CAMPAGNE
    cellule.font = Font(bold=True, size=15, color=BLANC)
    cellule.fill = PatternFill("solid", fgColor=BLEU)
    cellule.alignment = Alignment(horizontal="center", vertical="center")
    feuille.row_dimensions[1].height = 30

    cellule = feuille[_fusionner(feuille, 2)]
    cellule.value = (
        "SOCADEL, Société Camerounaise d'Electricité · opération NEXT LTD"
    )
    cellule.font = Font(size=9, color=GRIS_DOUX)
    cellule.alignment = Alignment(horizontal="center", vertical="center")
    feuille.row_dimensions[2].height = 18

    ligne = 3
    if nom_agent or date_travail:
        cellule = feuille[_fusionner(feuille, ligne)]
        cellule.value = f"Agent : {nom_agent or '.' * 30}    Date : {date_travail or '.' * 14}"
        cellule.font = Font(size=10, bold=True, color=GRIS_TEXTE)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        feuille.row_dimensions[ligne].height = 18
        ligne += 1

    _poser_logo(feuille)
    return ligne + 1


def _poser_logo(feuille: Worksheet) -> None:
    """Pose la marque en tête de feuille, si l'image est disponible.

    Excel n'a pas de filigrane : une image de fond se décalerait au premier
    ajout de ligne. Le logo est donc ancré en tête, où il s'imprime.
    """
    if not LOGO.exists():
        return
    try:
        image = ImageXl(str(LOGO))
        image.width, image.height = 118, 40
        image.anchor = "A1"
        feuille.add_image(image)
    except Exception:
        # Une image absente ou illisible ne doit pas priver l'agent de son
        # bordereau : le classeur part sans marque.
        pass


def _bloc(feuille: Worksheet, bloc: BlocTerrain, depart: int) -> int:
    """Bandeau de tournée, en-têtes de colonnes, puis la grille de saisie."""
    bordure = _bordure()
    total = len(bloc.lignes)

    # Bandeau, calqué sur la ligne 3 du classeur source.
    bandeau = ("ITINERAIRE", str(bloc.code), "Total client", str(total),
               "RAPPORT", "OK / MRA")
    for index, valeur in enumerate(bandeau, start=1):
        cellule = feuille.cell(row=depart, column=index, value=valeur)
        cellule.border = bordure
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        libelle = index in (1, 3, 5)
        cellule.font = Font(bold=True, size=9,
                            color=BLEU_SOMBRE if libelle else GRIS_TEXTE)
        if libelle:
            cellule.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
    feuille.row_dimensions[depart].height = 20

    if bloc.libelle:
        cellule = feuille[_fusionner(feuille, depart + 1)]
        cellule.value = f"Zone : {bloc.libelle}"
        cellule.font = Font(size=9, italic=True, color=GRIS_DOUX)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        depart += 1

    # En-têtes de colonnes.
    entete = depart + 1
    for index, (nom, _) in enumerate(COLONNES, start=1):
        cellule = feuille.cell(row=entete, column=index, value=nom)
        cellule.font = Font(bold=True, size=10, color=BLANC)
        cellule.fill = PatternFill("solid", fgColor=BLEU)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        cellule.border = bordure
    feuille.row_dimensions[entete].height = 22

    # Grille : quatre colonnes pré-remplies, deux laissées au stylo.
    premiere = entete + 1
    saisie = len(COLONNES) - 1
    for decalage, ligne_client in enumerate(bloc.lignes):
        rang = premiere + decalage
        valeurs = (
            ligne_client.ref_geo,
            ligne_client.numero_compteur,
            ligne_client.nom,
            ligne_client.service_no,
            "",
            "",
        )
        for index, valeur in enumerate(valeurs, start=1):
            cellule = feuille.cell(row=rang, column=index, value=valeur)
            cellule.border = bordure
            cellule.font = Font(size=10, color=GRIS_TEXTE)
            cellule.alignment = Alignment(
                horizontal="center" if index >= saisie else "left",
                vertical="center",
            )
        feuille.row_dimensions[rang].height = 19

    # Lignes vierges pour les clients hors référentiel.
    for decalage in range(LIGNES_VIERGES):
        rang = premiere + len(bloc.lignes) + decalage
        for index in range(1, len(COLONNES) + 1):
            cellule = feuille.cell(row=rang, column=index, value="")
            cellule.border = bordure
            cellule.fill = PatternFill("solid", fgColor=CREME)
        feuille.row_dimensions[rang].height = 19

    derniere = premiere + len(bloc.lignes) + LIGNES_VIERGES - 1
    _contraindre_rapport(feuille, premiere, derniere)
    return derniere + 2


def _contraindre_rapport(feuille: Worksheet, premiere: int, derniere: int) -> None:
    """Liste déroulante sur RAPPORT : deux valeurs, pas une de plus.

    C'est la source d'anomalie la plus fréquente à l'import, un « ok » ou un
    « valide » écrit à la main. La contrainte ne gêne pas l'agent, qui remplit
    de toute façon au stylo, mais protège le superviseur qui ressaisit.
    """
    colonne = get_column_letter(len(COLONNES) - 1)
    validation = DataValidation(
        type="list",
        formula1='"OK,MRA"',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle="Valeur non autorisée",
        error="Écrivez OK si le client est enrôlé, MRA s'il reste à relancer.",
    )
    feuille.add_data_validation(validation)
    validation.add(f"{colonne}{premiere}:{colonne}{derniere}")


def _legende(feuille: Worksheet, depart: int) -> int:
    """Ce que l'agent doit écrire, expliqué sur la feuille elle-même."""
    entrees = (
        ("OK", "Le client est allé au bout du parcours WhatsApp : il figure "
               "désormais dans la base."),
        ("MRA", "Le numéro est pris mais l'enrôlement reste à faire : il sera "
                "relancé depuis Gestion Campagnes, sur MRA."),
        ("N° WHATSAPP", "À renseigner quand le client est absent et qu'un "
                        "proche donne son numéro, ou quand le relevé diffère."),
    )
    bordure = _bordure()
    derniere = get_column_letter(len(COLONNES))

    for decalage, (cle, explication) in enumerate(entrees):
        rang = depart + decalage
        cellule = feuille.cell(row=rang, column=1, value=cle)
        cellule.font = Font(bold=True, size=9, color=BLEU_SOMBRE)
        cellule.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
        cellule.alignment = Alignment(horizontal="center", vertical="center")
        cellule.border = bordure

        feuille.merge_cells(f"B{rang}:{derniere}{rang}")
        cellule = feuille[f"B{rang}"]
        cellule.value = explication
        cellule.font = Font(size=9, color=GRIS_TEXTE)
        cellule.alignment = Alignment(horizontal="left", vertical="center")
        cellule.border = bordure
        feuille.row_dimensions[rang].height = 17

    return depart + len(entrees)


def _signature(feuille: Worksheet, depart: int) -> None:
    """Le pied du modèle : c'est lui qui authentifie le retour de tournée."""
    bordure = _bordure()
    derniere = get_column_letter(len(COLONNES))
    milieu = get_column_letter(len(COLONNES) // 2)
    suivante = get_column_letter(len(COLONNES) // 2 + 1)

    feuille.merge_cells(f"A{depart}:{derniere}{depart}")
    cellule = feuille[f"A{depart}"]
    cellule.value = "DATE ET SIGNATURE SUPERVISEUR SOCADEL / ENTREPRISE"
    cellule.font = Font(bold=True, size=10, color=BLEU_SOMBRE)
    cellule.fill = PatternFill("solid", fgColor=BLEU_CLAIR)
    cellule.alignment = Alignment(horizontal="center", vertical="center")
    cellule.border = bordure
    feuille.row_dimensions[depart].height = 22

    for rang, hauteur, valeurs in (
        (depart + 1, 18, ("Superviseur SOCADEL", "Agent de terrain")),
        (depart + 2, 50, ("", "")),
    ):
        feuille.merge_cells(f"A{rang}:{milieu}{rang}")
        feuille.merge_cells(f"{suivante}{rang}:{derniere}{rang}")
        for reference, valeur in ((f"A{rang}", valeurs[0]),
                                  (f"{suivante}{rang}", valeurs[1])):
            cellule = feuille[reference]
            cellule.value = valeur
            cellule.font = Font(bold=True, size=9, color=GRIS_DOUX)
            cellule.alignment = Alignment(horizontal="center", vertical="center")
            cellule.border = bordure
        feuille.row_dimensions[rang].height = hauteur
